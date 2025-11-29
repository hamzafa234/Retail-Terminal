import json
import os
import sys
import yfinance as yf
from datetime import datetime
import pandas as pd
import requests
from openai import OpenAI
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from flask import Flask, render_template, url_for, request, redirect
from flask_sqlalchemy import SQLAlchemy

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///test.db'
db = SQLAlchemy(app)

client = OpenAI(
  api_key="xxxxxxxxxxx"
)

class FinancialDataFetcher:
    def __init__(self, ticker):
        self.ticker = ticker.upper()
        self.stock = yf.Ticker(self.ticker)
        
    def convert_dataframe_to_dict(self, df):
        """Convert pandas DataFrame to dictionary with proper formatting"""
        if df is None or df.empty:
            return []
        
        # Convert to dictionary with dates as strings
        result = []
        for col in df.columns:
            period_data = {'date': str(col.date())}
            for idx in df.index:
                value = df.loc[idx, col]
                # Convert numpy/pandas types to Python native types
                if pd.isna(value):
                    period_data[idx] = None
                elif isinstance(value, (pd.Timestamp, datetime)):
                    period_data[idx] = str(value)
                else:
                    try:
                        period_data[idx] = float(value)
                    except (ValueError, TypeError):
                        period_data[idx] = str(value)
            result.append(period_data)
        
        return result
    
    def get_balance_sheet(self):
        """Fetch most recent quarterly balance sheet data"""
        quarterly_bs = self.stock.quarterly_balance_sheet
        
        if quarterly_bs is None or quarterly_bs.empty:
            return []
        
        # Get only the most recent quarter (first column)
        most_recent = quarterly_bs.iloc[:, [0]]
        return self.convert_dataframe_to_dict(most_recent)
    
    def get_income_statement(self):
        """Fetch income statement data"""
        return self.convert_dataframe_to_dict(self.stock.financials)
    
    def get_cash_flow(self):
        """Fetch cash flow statement data"""
        return self.convert_dataframe_to_dict(self.stock.cashflow)
    
    def get_all_financials(self):
        """Fetch all financial statements and return as dictionary"""
        print(f"\nFetching financial data for {self.ticker}...")
        
        try:
            # Get basic info to verify ticker exists
            info = self.stock.info
            company_name = info.get('longName', self.ticker)
            
            financials = {
                'ticker': self.ticker,
                'companyName': company_name,
                'currency': info.get('currency', 'USD'),
                'fetchDate': datetime.now().isoformat(),
                'balanceSheet': self.get_balance_sheet(),
                'incomeStatement': self.get_income_statement(),
                'cashFlowStatement': self.get_cash_flow()
            }
            
            return financials
            
        except Exception as e:
            print(f"Error fetching data: {e}")
            return None
    
    def save_to_json(self, filename=None):
        """Fetch financial data and save to JSON file"""
        financials = self.get_all_financials()
        
        if not financials:
            print("Failed to fetch financial data")
            return False
        
        if filename is None:
            filename = f"{self.ticker}_financials.json"
        
        try:
            with open(filename, 'w') as f:
                json.dump(financials, f, indent=2)
            
            # Print summary of data retrieved
            bs_periods = len(financials['balanceSheet'])
            is_periods = len(financials['incomeStatement'])
            cf_periods = len(financials['cashFlowStatement'])
            
            print(f"\nData Retrieved:")
            print(f"  Balance Sheet: {bs_periods} quarterly period (most recent)")
            print(f"  Income Statement: {is_periods} annual periods")
            print(f"  Cash Flow: {cf_periods} annual periods")
            
            return True
            
        except Exception as e:
            print(f"Error saving to JSON: {e}")
            return False

def get_10y_treasury_yield():
    """
    Fetches the latest 10-Year US Treasury yield (^TNX) from Yahoo Finance.
    Returns the yield as a float (in percent).
    """
    ticker = yf.Ticker("^TNX")
    data = ticker.history(period="1d")
    
    if data.empty:
        raise ValueError("No data returned for ^TNX.")
    
    latest_yield = data["Close"].iloc[-1]
    return latest_yield

def get_20y_treasury_yield():
    """
    Fetches the latest 20-Year US Treasury yield (^TYX) from Yahoo Finance.
    Returns the yield as a float (in percent).
    """
    ticker = yf.Ticker("^TYX")  # 20-Year Treasury Index
    data = ticker.history(period="1d")
    
    if data.empty:
        raise ValueError("No data returned for ^TYX.")
    
    latest_yield = data["Close"].iloc[-1]
    return latest_yield

def get_30y_treasury_yield():
    """
    Fetches the latest 30-Year US Treasury yield (^TYX) from Yahoo Finance.
    Returns the yield as a float (in percent).
    """
    ticker = yf.Ticker("^TYX")  # 30-Year Treasury Index
    data = ticker.history(period="1d")
    
    if data.empty:
        raise ValueError("No data returned for ^TYX.")
    
    latest_yield = data["Close"].iloc[-1]
    return latest_yield


def credit_spread_analysis(ticker):

    try: 
        with open(f'{ticker}_financials.json', 'r') as file:
            data = json.load(file)
    except:
        main(ticker)
        with open(f'{ticker}_financials.json', 'r') as file:
            data = json.load(file)

    response = client.responses.create(
        model="gpt-5-nano",
        tools=[{"type": "web_search"}],
        input="get me the yield to maturity for" + ticker + "that is due in more than 5 years. Chose the bond that was traded the most recently. Do not ask any follow up questions. Return a list that has two floats. Do not return any text other than the list. If the yeild is 5.49 percent and has a maturity date of 2065 you will return [0.0549, 2065.0]. Do not return a link ever if the company has no bonds return [0.0, 0.0]",
        store=True,
    )

    bond_data = eval(response.output_text)
    yeild = bond_data[0]
    bond_maturity = bond_data[1]
    yeild = yeild * 100

    delta = bond_maturity - datetime.now().year
    if delta > 25:
        treasury_yield = get_30y_treasury_yield()
        tbill = 30
    elif delta > 15:
        treasury_yield = get_20y_treasury_yield()
        tbill = 20
    else:
        treasury_yield = get_10y_treasury_yield()
        tbill = 10
    
    return(bond_maturity, yeild, treasury_yield, tbill)

def compare(ticker):

    try: 
        with open(f'{ticker}_financials.json', 'r') as file:
            data = json.load(file)
    except:
        main(ticker)
        with open(f'{ticker}_financials.json', 'r') as file:
            data = json.load(file)

    response = client.responses.create(
        model="gpt-5-nano",
        tools=[{"type": "web_search"}],
        input="get me the main competitors for " + ticker + " Do not ask any follow up questions. Return a list of ticker symbols only. Do not return any text other than the list. only return up to 5 competitors. for example if the main competeitors are apple, microsoft, and google, you will return ['AAPL', 'MSFT', 'GOOGL']. If there are no competitors return an empty list [] do not return a link ever",
        store=True,
    )
    comp_data = eval(response.output_text)
    comp_data.append(ticker)
    lis = []
    for comp in comp_data:
        if comp is not None:
            temp = getCompVal(comp)
            lis.append(temp)
        
        


def format_large_number(num):
    """Format large numbers with B (billions), M (millions), or K (thousands)"""
    if num >= 1_000_000_000_000:  # Trillions
        return f"{num / 1_000_000_000_000:.2f}T"
    elif num >= 1_000_000_000:  # Billions
        return f"{num / 1_000_000_000:.2f}B"
    elif num >= 1_000_000:  # Millions
        return f"{num / 1_000_000:.2f}M"
    elif num >= 1_000:  # Thousands
        return f"{num / 1_000:.2f}K"
    else:
        return f"{num:.2f}"

def getCompVal(ticker):

    stock = yf.Ticker(ticker)
    cap = stock.info.get('marketCap', None)
    pe = stock.info.get('trailingPE', None)
    fpe = stock.info.get('forwardPE', None)
    y = stock.info.get('dividendYield', None)
    ev = stock.info.get('enterpriseValue', None)
    
    # Calculate Price to FCF manually
    price_to_fcf = None
    try:
        # Get free cash flow from cash flow statement
        cash_flow = stock.cashflow
        if not cash_flow.empty and 'Free Cash Flow' in cash_flow.index:
            # Get the most recent free cash flow (first column)
            fcf = cash_flow.loc['Free Cash Flow'].iloc[0]
            
            # If FCF is positive and we have market cap, calculate ratio
            if fcf and fcf > 0 and cap:
                price_to_fcf = cap / fcf
    except Exception as e:
        # If calculation fails, leave as None
        pass

    list = [cap, ev, pe, fpe, y, price_to_fcf]
    return list  

def capital_structure_summary(ticker):

    try: 
        with open(f'{ticker}_financials.json', 'r') as file:
            data = json.load(file)
    except:
        main(ticker)
        with open(f'{ticker}_financials.json', 'r') as file:
            data = json.load(file)

    stock = yf.Ticker(ticker)
    cap = stock.info.get('marketCap', None)

    total_debt = data['balanceSheet'][0]['Total Debt']
    latest_balance_sheet = data['balanceSheet'][0]
    income = data['incomeStatement'][0]
    total_debt_formatted = format_large_number(total_debt)
    # Use .get() method to safely retrieve preferred stock value
    pre = data['balanceSheet'][0].get('Preferred Stock Equity', 0) or 0
    cash = latest_balance_sheet['Cash And Cash Equivalents']
    cash_formatted = format_large_number(cash)
    EV = stock.info.get('marketCap', None) + total_debt - latest_balance_sheet['Cash And Cash Equivalents'] + pre

    shares_outstanding = income['Basic Average Shares']
    shares_outstanding_formatted = format_large_number(shares_outstanding)
    netdebtpershare = (total_debt - latest_balance_sheet['Cash And Cash Equivalents']) / income['Basic Average Shares']
    if pre != 0:
        preferred_formatted = format_large_number(pre)
    else:
        preferred_formatted = 0         
    return([preferred_formatted, netdebtpershare, shares_outstanding_formatted, total_debt_formatted, format_large_number(cap), cash_formatted, format_large_number(EV)])

def calculate_growth_rate(terminal_value, last_fcf, wacc):
    # Rearranged formula: g = (TV × WACC - FCF) / (TV + FCF)
    numerator = (terminal_value * wacc) - last_fcf
    denominator = terminal_value + last_fcf
    growth_rate = numerator / denominator
    
    return growth_rate

def generate_excel(ticker, type):

    try: 
        with open(f'{ticker}_financials.json', 'r') as file:
            data = json.load(file)
    except:
        main(ticker)
        with open(f'{ticker}_financials.json', 'r') as file:
            data = json.load(file)

    income_statement = data['incomeStatement'][0]
    cash_flow_statement = data['cashFlowStatement'][0]
    balance_sheet = data['balanceSheet'][0]

    wb = openpyxl.load_workbook('template.xlsx')
    ws = wb.active
        
    debt = balance_sheet['Total Debt']
    cash = balance_sheet['Cash And Cash Equivalents']
    shares = income_statement['Basic Average Shares']

    ws["U17"] = debt/1000000
    ws["U15"] = cash/1000000
    ws["U21"] = shares/1000000

    rev = income_statement['Total Revenue']
    cogs = income_statement['Cost Of Revenue']
    operating_expense = income_statement['Operating Expense']
    other_expense = income_statement['Net Interest Income']
    tax = income_statement['Tax Provision']

    # Use .get() method with default value of 0 to handle None or missing keys
    div = income_statement.get("Preferred Stock Dividends", 0) or 0
    div += income_statement.get('Otherunder Preferred Stock Dividend', 0) or 0

    ws["C2"] = rev/1000000
    ws["C4"] = cogs/1000000
    ws["C10"] = operating_expense/1000000
    ws["C16"] = other_expense/1000000
    ws["C21"] = tax/1000000
    ws["C25"] = div/1000000

    lis = wacc_no_print(ticker)    
    wacc = lis[0]
    cost_of_equity = lis[1]
    cost_of_debt = lis[2]
    weight_of_equity = lis[3]
    weight_of_debt = lis[4]

    ws["U10"] = wacc/100
    ws["U8"] = weight_of_equity
    ws["U6"] = weight_of_debt
    ws["U2"] = cost_of_equity/100
    ws["U4"] = cost_of_debt/100

    ws["L6"] = cash_flow_statement['Change In Working Capital']/1000000
    ws["L10"] = cash_flow_statement['Capital Expenditure']/1000000

    ws["L12"] = cash_flow_statement['Free Cash Flow']/1000000 - div/1000000

    ws["L4"] = cash_flow_statement['Operating Cash Flow']/1000000 - cash_flow_statement['Change In Working Capital']/1000000 - income_statement['Net Income']/1000000

    ws["Q16"] = discount_factor(wacc/100, 5)
    ws["P16"] = discount_factor(wacc/100, 4)
    ws["O16"] = discount_factor(wacc/100, 3)
    ws["N16"] = discount_factor(wacc/100, 2)
    ws["M16"] = discount_factor(wacc/100, 1)

    # Save the workbook    
    if (type == "default"):
        wb.save(f'{ticker}_financial_model.xlsx')
        return
    elif(type == "expectation"):
        stock = yf.Ticker(ticker)
        price_to_fcf = None
        cap = stock.info.get('marketCap', None)
        marketcap = cap
    
        # Get free cash flow from cash flow statement
        cash_flow = stock.cashflow
        if not cash_flow.empty and 'Free Cash Flow' in cash_flow.index:
            # Get the most recent free cash flow (first column)
            fcf = cash_flow.loc['Free Cash Flow'].iloc[0]
            
            # If FCF is positive and we have market cap, calculate ratio
            if fcf and fcf > 0 and cap:
                price_to_fcf = cap / fcf

        ws["M18"] = (fcf * ws["M16"].value)/1000000

        yearone = (price_to_fcf/200) + 1
        ws["D3"] = yearone
        ws["D5"] = yearone
        ws["D11"] = yearone
        ws["C22"] = tax/income_statement['Pretax Income']
        tax_rate = tax/income_statement['Pretax Income']
        ws["D22"] = tax_rate
        ws["E22"] = tax_rate
        ws["F22"] = tax_rate
        ws["G22"] = tax_rate
        ws["H22"] = tax_rate
        ws["M5"] = yearone
        ws["M7"] = yearone
        ws["M11"] = yearone

        ws["D3"].number_format = '0.00'
        ws["D5"].number_format = '0.00'
        ws["D11"].number_format = '0.00'
        ws["M5"].number_format = '0.00'
        ws["M7"].number_format = '0.00'
        ws["M11"].number_format = '0.00'

        fcf = fcf * yearone 

        ws["N18"] = (fcf * ws["N16"].value)/1000000

        cap = cap * (1 + wacc/100)
        pfcftwo = cap / fcf
        yeartwo = (pfcftwo/200) + 1
        ws["E3"] = yeartwo
        ws["E5"] = yeartwo
        ws["E11"] = yeartwo
        ws["N5"] = yeartwo
        ws["N7"] = yeartwo
        ws["N11"] = yeartwo

        ws["E3"].number_format = '0.00'
        ws["E5"].number_format = '0.00'
        ws["E11"].number_format = '0.00'
        ws["N5"].number_format = '0.00'
        ws["N7"].number_format = '0.00'
        ws["N11"].number_format = '0.00'

        fcf = fcf * yeartwo

        ws["O18"] = (fcf * ws["O16"].value)/1000000

        cap = cap * (1 + wacc/100)
        pfcfthree = cap / fcf
        yearthree = (pfcfthree/200) + 1
        ws["F3"] = yearthree
        ws["F5"] = yearthree
        ws["F11"] = yearthree
        ws["O5"] = yearthree
        ws["O7"] = yearthree
        ws["O11"] = yearthree

        ws["F3"].number_format = '0.00'
        ws["F5"].number_format = '0.00'
        ws["F11"].number_format = '0.00'
        ws["O5"].number_format = '0.00'
        ws["O7"].number_format = '0.00'
        ws["O11"].number_format = '0.00'

        fcf = fcf * yearthree
        ws["P18"] = (fcf * ws["P16"].value)/1000000
        cap = cap * (1 + wacc/100)
        pfcffour = cap / fcf
        yearfour = (pfcffour/200) + 1
        ws["G3"] = yearfour
        ws["G5"] = yearfour
        ws["G11"] = yearfour
        ws["P5"] = yearfour
        ws["P7"] = yearfour
        ws["P11"] = yearfour

        ws["P11"].number_format = '0.00'
        ws["G3"].number_format = '0.00'
        ws["G5"].number_format = '0.00'
        ws["G11"].number_format = '0.00'
        ws["P5"].number_format = '0.00'
        ws["P7"].number_format = '0.00'

        fcf = fcf * yearfour
        ws["Q18"] = (fcf * ws["Q16"].value)/1000000
        cap = cap * (1 + wacc/100)
        pfcffive = cap / fcf
        yearfive = (pfcffive/200) + 1
        ws["H3"] = yearfive
        ws["H5"] = yearfive
        ws["H11"] = yearfive
        ws["Q5"] = yearfive
        ws["Q7"] = yearfive
        ws["Q11"] = yearfive

        ws["H3"].number_format = '0.00'
        ws["H5"].number_format = '0.00'
        ws["H11"].number_format = '0.00'
        ws["Q5"].number_format = '0.00'
        ws["Q7"].number_format = '0.00'
        ws["Q11"].number_format = '0.00'

        # subtracting too much need to account for quaters that have already happened
        TV = (marketcap/1000000) - (ws["M18"].value + ws["N18"].value + ws["O18"].value + ws["P18"].value + ws["Q18"].value) - balance_sheet['Cash And Cash Equivalents']/1000000 + balance_sheet['Total Debt']/1000000
        ws["Q20"] = TV

        TTV = TV 

        TV = TV*(1 + wacc/100)**5

        ws["K24"] = TV
        
        g = calculate_growth_rate(TV*1000000, fcf, wacc/100)

        ws["J25"] = g

        wb.save(f'{ticker}_financial_model.xlsx')
        return

def discount_factor(rate, period):
    return 1 / ((1 + rate) ** period)

def wacc_no_print(ticker):
    global yield_10y
    yield_10y = get_10y_treasury_yield()

    stock = yf.Ticker(ticker)
    cap = stock.info.get('marketCap', None)
    beta = stock.info.get("beta")

    ERP = 10 - yield_10y

    with open(f'{ticker}_financials.json', 'r') as file:
            data = json.load(file)

    # Access the first balance sheet entry
    latest_balance_sheet = data['balanceSheet'][0]

    income = data['incomeStatement'][0]

    cash = data['cashFlowStatement'][0]

    # Get Total Debt
    total_debt = latest_balance_sheet['Total Debt']

    tax_rate = income["Tax Rate For Calcs"]


    weight_of_debt = total_debt / (total_debt + cap)
    weight_of_equity = cap / (total_debt + cap)

    weight_of_debt = f"{weight_of_debt:.4f}"
    weight_of_equity = f"{weight_of_equity:.4f}"

    if 'Interest Expense' in income:
        interest_expense = income['Interest Expense']
    else:
        interest_expense = cash['Interest Paid Supplemental Data']
    
    if(interest_expense == None):
        cost_of_debt = 0
    else:
        cost_of_debt = abs(interest_expense / total_debt) * 100

    cost_of_equity = yield_10y + beta * ERP
    
    WACC = (float(weight_of_equity) * cost_of_equity) + (float(weight_of_debt) * cost_of_debt * (1 - tax_rate))

    lis = [WACC, cost_of_equity, cost_of_debt, weight_of_equity, weight_of_debt]

    return lis

def main(command):

    global ticker
    ticker = command if command != "" else None
    
    fetcher = FinancialDataFetcher(ticker)
    success = fetcher.save_to_json()

    stock = yf.Ticker(ticker)
    global beta 
    global cap 
    global current_price
    info = stock.info
    beta = info.get("beta")
    cap = info.get('marketCap', None)
    global current_price
    current_price = stock.info.get("currentPrice")

if __name__ == "__main__":
    with app.app_context():
        db.create_all()  
    app.run(debug=True)